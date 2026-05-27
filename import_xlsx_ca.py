#!/usr/bin/env python3
"""
Importa os dois Excel (principal + backup) da conta CA - Pré-Campanha 2026
e atualiza snapshots_diarios no Supabase.
"""

import os
import logging
from collections import defaultdict
import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

ACCOUNT_ID   = "act_4085311614896655"
ACCOUNT_NAME = "CA - Pré- Campanha 2026"

PRINCIPAL = "/Users/alexsandro/Downloads/Relatório-sem-título-jan-1-2026-a-mai-15-2026_principal.xlsx"
BACKUP    = "/Users/alexsandro/Downloads/Relatório-sem-título-jan-1-2026-a-mai-15-2026_Backup.xlsx"


def f(v):
    try:
        return float(v or 0)
    except:
        return 0.0


def aggregate_principal(path):
    """
    Cols: 0=campanha 1=adset 2=dia 3=alcance 4=impressoes 8=spend
          11=cliques_link 21=cliques_todos
    Row 1 é totais globais (sem dia) — pular.
    """
    days = defaultdict(lambda: dict(spend=0.0, impressoes=0, alcance=0, cliques=0))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        dia = row[2]
        if not dia or not row[0]:  # pula linha de totais
            continue
        date = str(dia)[:10]
        days[date]["spend"]     += f(row[8])
        days[date]["impressoes"] += int(f(row[4]))
        days[date]["alcance"]    += int(f(row[3]))
        days[date]["cliques"]    += int(f(row[11]))  # Cliques no link
    return days


def aggregate_backup(path):
    """
    Cols: 0=campanha 1=adset 2=dia 5=alcance 6=impressoes 11=spend
          15=cliques_link
    """
    days = defaultdict(lambda: dict(spend=0.0, impressoes=0, alcance=0, cliques=0))
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        dia = row[2]
        if not dia:
            continue
        date = str(dia)[:10]
        days[date]["spend"]      += f(row[11])
        days[date]["impressoes"]  += int(f(row[6]))
        days[date]["alcance"]     += int(f(row[5]))
        days[date]["cliques"]     += int(f(row[15]))  # Cliques no link
    return days


class SupabaseClient:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.h = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        }

    def upsert(self, table, data):
        r = requests.post(
            f"{self.base}/{table}",
            headers={**self.h, "Prefer": "resolution=merge-duplicates,return=minimal"},
            json=data, timeout=15,
        )
        if r.status_code not in (200, 201):
            logger.error(f"Upsert falhou: {r.status_code} {r.text[:300]}")
            return False
        return True


def main():
    supabase_url = os.getenv("SUPABASE_URL")
    service_key  = os.getenv("SUPABASE_SERVICE_KEY")
    if not supabase_url or not service_key:
        logger.error("SUPABASE_URL ou SUPABASE_SERVICE_KEY não configurados")
        return

    logger.info("Lendo PRINCIPAL…")
    principal = aggregate_principal(PRINCIPAL)
    logger.info(f"  {len(principal)} dias")

    logger.info("Lendo BACKUP…")
    backup = aggregate_backup(BACKUP)
    logger.info(f"  {len(backup)} dias")

    # Merge: soma backup no principal
    merged = dict(principal)
    for date, bv in backup.items():
        if date in merged:
            merged[date]["spend"]      += bv["spend"]
            merged[date]["impressoes"] += bv["impressoes"]
            merged[date]["alcance"]    += bv["alcance"]
            merged[date]["cliques"]    += bv["cliques"]
        else:
            merged[date] = dict(bv)

    logger.info(f"Total de dias combinados: {len(merged)}")
    total_spend = sum(v["spend"] for v in merged.values())
    logger.info(f"Spend total combinado: R$ {total_spend:,.2f}")

    db = SupabaseClient(supabase_url, service_key)
    inserted = 0

    for date in sorted(merged.keys()):
        m = merged[date]
        spend    = round(m["spend"], 2)
        cliques  = m["cliques"]
        impressoes = m["impressoes"]
        alcance  = m["alcance"]
        leads    = 0  # arquivos não têm coluna de leads

        eq   = cliques if cliques > 0 else 0
        cpee = round(spend / eq, 4) if eq > 0 else 0.0
        ctr  = round(cliques / impressoes * 100, 4) if impressoes > 0 and cliques > 0 else 0.0
        cpc  = round(spend / cliques, 2) if cliques > 0 else 0.0

        row = {
            "data":        date,
            "account_id":  ACCOUNT_ID,
            "nome":        ACCOUNT_NAME,
            "spend":       spend,
            "impressoes":  impressoes,
            "alcance":     alcance,
            "cliques":     cliques,
            "eq":          eq,
            "leads":       leads,
            "cpee":        cpee,
            "ctr":         ctr,
            "cpc":         cpc,
        }

        if db.upsert("snapshots_diarios", row):
            logger.info(f"  {date}: spend=R${spend:,.2f} cliques={cliques} cpee={cpee}")
            inserted += 1

    logger.info(f"\nFinalizado: {inserted} dias atualizados")


if __name__ == "__main__":
    main()
